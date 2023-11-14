import glob
import os
import xlsxwriter
import openpyxl

first_output_file = 'raw_output.txt'
second_output_file = 'processed_output_1.txt'
classes = ['Annelids', 'Arthropods', 'Cnidarians', 'Echinoderms', 'fish', 'Mollusca', 'other-invertebrates', 'Porifera', 'unidentified-biology']
sourceVid = ""
excelName = "detections.xlsx"
worksheetName = "detections_wksht"
output_path_log = "output_path_log.txt"

def fixFormat(): # need to call this at the end to shift all cells up one row, by default processing multiple videos causes a 1 row buffer per video
  wb = openpyxl.load_workbook(excelName)
  ws = wb.active
  newData = []
  for row in ws.iter_rows(min_row=0, values_only=True): #rooting out any empty cells
    if row[0] != None and row[1] != None:
      obj = {"vid": row[0], "frame": row[1], "class": row[2], "x_left": row[3], "x_right": row[4], "y_up": row[5], "y_low": row[6]}
      newData.append(obj)
  os.remove(excelName)
  newBook = xlsxwriter.Workbook(excelName)
  newWorksheet = newBook.add_worksheet(worksheetName)
  bold = newBook.add_format({'bold': True})
  for index, entry in enumerate(newData): #making necesary augmentations (the replaced classes needed to be one token long before excel)
        if entry["class"] == classes[4]:
            newWorksheet.write(index, 2, "Vertebrates: Fishes")
        elif entry["class"] == classes[6]:
            newWorksheet.write(index, 2, "Other Invertebrates")
        elif entry["class"] == classes[8]:
            newWorksheet.write(index, 2, "Unidentified Biology")
        else:
            newWorksheet.write(index, 2, entry["class"])
        newWorksheet.write(index, 0, entry["vid"])
        newWorksheet.write(index, 1, entry["frame"])
        newWorksheet.write(index, 3, entry["x_left"])
        newWorksheet.write(index, 4, entry["x_right"])
        newWorksheet.write(index, 5, entry["y_up"])
        newWorksheet.write(index, 6, entry["y_low"])
  newWorksheet.set_row(0, None, bold)
  newBook.close()

def parseFrame(title):
    frameStr = ""
    dotHit = False
    finished = False
    #parsing into reversed string of numbers
    for i in range(len(title) - 1, -1, -1):
        if finished == False:
            if dotHit == False:
                if title[i] == '.':
                    dotHit = True
            elif title[i] == '_':
                finished = True
            elif dotHit:
                frameStr += title[i]
        else:
            break
    #reversing the numbers string back
    fs2 = ""
    for i in range(len(frameStr) - 1, -1, -1):
        fs2 += frameStr[i]
    return fs2

def logOutputPath(outputPath): #logs yolo's output path to a txt file so we can go through each folder, grab the video and delete it - this file will be deleted when not needed
   if os.path.exists(output_path_log):
      with open(output_path_log, 'a') as out:
         out.write(outputPath + "\n")
   else:
      with open(output_path_log, 'w') as opl:
         opl.write(outputPath + "\n")

def determineOutputPath(): 
    current_dir = os.path.dirname(os.path.abspath(__file__))
    subfolder = os.path.join(current_dir, 'runs/detect/')
    max = 0
    latest = ""
    dir_list = os.listdir(subfolder)
    for file in dir_list:
        if len(file) > 3:
            expNum = int(file[3:len(file)])
            if expNum > max:
                max = expNum
    if max == 0:
        latest = "exp"
    else:
        latest = "exp" + str(max)
    addy = "./runs/detect/" + latest + "/labels/"
    print(addy)
    logOutputPath("./runs/detect/" + latest)
    return addy

def get_class(index): # this accounts for the use of models with different classes so the program can run on different models without an error - if something is classified as an invalid class, we log it as unidentified biology
   if index >= len(classes):
      return "Unidentified-biology" + str(index)
   else:
      return classes[index]
   
with open("sourceVid.txt", 'r') as sv:
  sourceVid = sv.readline()
txts_path = determineOutputPath()
with open(first_output_file , 'w') as f_out:
    for txt_file in glob.glob(txts_path + '*.txt'):
        frame = parseFrame(os.path.basename(txt_file)) #parses the frame number from the title of the file
        with open(txt_file, 'r') as f_in:
            lines = f_in.readlines()
            for line in lines:
                if line != "" and line != " ":
                    f_out.write(sourceVid + " " + frame + " " + line)
with open(second_output_file, 'w') as ff_out:
    with open(first_output_file, 'r') as pf_in:
      lines = pf_in.readlines()
      for line in lines:
        tokens = line.split()
        source = tokens[0][7:len(tokens[0])]
        frame = str(int(tokens[1]))
        animal = get_class(int(tokens[2])) 
        x_bound_left = str(tokens[3]) #x1
        y_bound_top = str(tokens[4]) #y1
        x_bound_right = str(float(tokens[5]) + float(tokens[3])) #x2
        y_bound_bottom = str(float(tokens[6]) + float(tokens[4])) #y2
        ff_out.write(source + " " + frame + " " + animal + " " + x_bound_left + " " + x_bound_right + " " + y_bound_top + " " + y_bound_bottom + "\n")
#getting data in list of dictionaries to be sorted
data = []
excelName = "detections.xlsx"
worksheetName = "detections_wksht"
with open(second_output_file, 'r') as sf_in:
    lines = sf_in.readlines()
    for line in lines:
        dict = {}
        tokens = line.split()
        dict['vid'] = tokens[0]
        dict['frame'] = tokens[1]
        dict['class'] = tokens[2]
        dict['x_left'] = tokens[3]
        dict['x_right'] = tokens[4]
        dict['y_up'] =  tokens[5]
        dict['y_low'] = tokens[6]
        data.append(dict)
sortedData = sorted(data, key=lambda k: int(k['frame'])) #sorting data by frame 
# write to excel from list of dictionaries
if os.path.exists(excelName): # if this is not the first video we're processing in this cycle
    # read in all the data from this into a list of dictionaries
    # delete this excel file, having stored all the data in the dicts
    # write all of them to a new excel file with the same name 
    prevData = []
    book = openpyxl.load_workbook(excelName)
    sheet = book.active
    for row in sheet.iter_rows(values_only=True):
        if row[0] != None:
          obj = {"vid": row[0], "frame": row[1], "class": row[2], "x_left": row[3], "x_right": row[4], "y_up": row[5], "y_low": row[6]}
          prevData.append(obj) #the whole row of data
        else:
          print("GOT")
    os.remove(excelName) # deleting excel file
    prevData.extend(sortedData) #adding new data to the rest
    workbook = xlsxwriter.Workbook(excelName)
    worksheet = workbook.add_worksheet(worksheetName)
    for index, entry in enumerate(prevData): #writing all of it to the same excel worksheet
        if entry["vid"] != None: #double checking
          worksheet.write(index+1, 0, entry["vid"])
          worksheet.write(index+1, 1, entry["frame"])
          worksheet.write(index+1, 2, entry["class"])
          worksheet.write(index+1, 3, entry["x_left"])
          worksheet.write(index+1, 4, entry["x_right"])
          worksheet.write(index+1, 5, entry["y_up"])
          worksheet.write(index+1, 6, entry["y_low"])
    workbook.close()
else: # if this is the first video we are processing in this cycle
    workbook = xlsxwriter.Workbook(excelName)
    worksheet = workbook.add_worksheet(worksheetName)
    worksheet.write(0,0,"Source Video")
    worksheet.write(0,1,"Current Frame")
    worksheet.write(0,2,"Classification")
    worksheet.write(0,3,"X Bound, Left")
    worksheet.write(0,4,"X Bound, Right")
    worksheet.write(0,5,"Y Bound, Upper")
    worksheet.write(0,6,"Y Bound, Lower")
    for index, entry in enumerate(sortedData):
        worksheet.write(index+1, 0, entry["vid"])
        worksheet.write(index+1, 1, entry["frame"])
        worksheet.write(index+1, 2, entry["class"])
        worksheet.write(index+1, 3, entry["x_left"])
        worksheet.write(index+1, 4, entry["x_right"])
        worksheet.write(index+1, 5, entry["y_up"])
        worksheet.write(index+1, 6, entry["y_low"])
    workbook.close()
fixFormat()
# delete all created files other than the spreadsheet & video
os.remove(first_output_file)
os.remove(second_output_file)
os.remove("sourceVid.txt")
for file in os.listdir(txts_path):
    file_name = os.path.join(txts_path, file)
    os.remove(file_name)
os.rmdir(txts_path)